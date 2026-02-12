import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter
import datetime
import os
import string
from openpyxl.formatting import Rule
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border,Side
from collections import defaultdict
import json
import sys
import warnings
warnings.filterwarnings("ignore")

def exact_clone_well_info(reffile) :
    merged_result_info = [["well","gene_id","clone_id","barcode_info"]]

    ref_wb = load_workbook(reffile)
    for sheet in ref_wb.sheetnames :
        ws = ref_wb[sheet]
        for iter_row in range(4,ws.max_row+1) :
            if ws.cell(iter_row,1).value :
                tmp = []
                for iter_col in [1,2,9,10] :
                    tmp.append(ws.cell(iter_row,iter_col).value)
                merged_result_info.append(tmp)

    result = pd.DataFrame(merged_result_info[1:],columns=merged_result_info[0])
    return result

def add_column_Choose_orNot(indata):
    indata["Choosed"] = ""
    _tmp_filter = indata.loc[(indata["Nanopore_Result"]=="Right") | (indata["Nanopore_Result"]=="Double_Peak")]
    indata.loc[_tmp_filter.sort_values(["Nanopore_Result"],ascending=False).groupby(["gene_id"]).head(2).index,"Choosed"] = "Choose"

    return indata

def add_column_RightPer(_deal_data) :
    keep_columns = list(_deal_data)
    _deal_data["Nanopore_right_1"] = _deal_data.groupby(["gene_id"])["Nanopore_Result"].transform(lambda x: (x=='Right').sum())
    _deal_data["Nanopore_right_2"] = _deal_data.groupby(["gene_id"])["Nanopore_Result"].transform(lambda x: (x=='Double_Peak').sum())
    _deal_data["Nanopore_right_3"] = _deal_data["Nanopore_right_1"] + _deal_data["Nanopore_right_2"]

    tmp = _deal_data.groupby(["gene_id","Nanopore_right_1","Nanopore_right_3"])["clone_id"].count().reset_index()
    tmp["Nanopore_right_noDP"] = " " + tmp["Nanopore_right_1"].astype(str) + "/" + tmp["clone_id"].astype(str)
    tmp["Nanopore_right_hasDP"] = " " + tmp["Nanopore_right_3"].astype(str) + "/" + tmp["clone_id"].astype(str)
    tmp = tmp[["gene_id","Nanopore_right_noDP","Nanopore_right_hasDP"]]

    result = pd.merge(tmp,_deal_data,how="outer",on="gene_id")
    
    keep_columns.append("Nanopore_right_noDP")
    keep_columns.append("Nanopore_right_hasDP")
    return result[keep_columns]


def extra_snp_info(instr: str) -> list :
    result_list = []
    for tmp_str in re.split('[(,\')\[\]\s+]',instr) :
        if tmp_str :
            result_list.append(tmp_str)
    return result_list

def extra_snp_info_pos(inlist: list, type: str) -> list :
    result_list = []
    type_1 = ["mutations","nucleotides","indels"]
    type_2 = ["variants","risks","risks2"]
    type_3 = ["deletions"]

    if type in type_1 :
        for _ in range(0,len(inlist),2) :
            result_list.append(inlist[_])
    if type in type_2 :
        for _ in inlist :
            result_list.append(_)
    if type in type_3 :
        for _ in inlist :
            if "-" in str(_) :
                for _pos in range(int(_.split("-")[0]),int(_.split("-")[1])+1):
                    result_list.append(str(_pos))
            else :
                result_list.append(str(_))
    return result_list

def extra_cloumn_info(indata: pd.DataFrame, type: str) -> list :
    result_list = []
    list_size = len(indata)
    todeal_list = indata[type].to_list()
    
    for i in range(list_size) :
        tmp_list = extra_snp_info_pos(extra_snp_info(todeal_list[i]),type)
        result_list.append(tmp_list)
    
    return result_list

def add_column_mutation_list(indata):
    # result_mutation_list = []
    # type_list = ["mutations","nucleotides","deletions","indels","variants"]
    result_mutation_list = extra_cloumn_info(indata,"mutations")
    type_list_2 = ["nucleotides","deletions","indels","variants"]

    for tmp_type in type_list_2 :
        tmp_result_list = extra_cloumn_info(indata,tmp_type)
        for _ in range(len(tmp_result_list)) :
            for _info in tmp_result_list[_] :
                result_mutation_list[_].append(_info)

    result_mutation_pos_list = []
    for _ in result_mutation_list :
        tmp_list = [re.sub(u"\D","",i) for i in _ ]
        tmp_list_2 = sorted(set(map(int,tmp_list)))
        result_mutation_pos_list.append(tmp_list_2)

    choose_xf_list = []
    for tmp_set in result_mutation_pos_list :
        count = 1
        if len(tmp_set) < 2 :
            count = len(tmp_set)
        else:
            start_pos = int(tmp_set[0])
            for j in range(1,len(tmp_set)):
                if int(tmp_set[j]) - int(start_pos) > 30 :
                    count = count + 1
                    start_pos = tmp_set[j]
        choose_xf_list.append(count)
    return result_mutation_list,choose_xf_list

def add_column_Double_Peak(indata):
    result_double_list = []
    result_max_double_ratio = []
    result_min_double_sum = []
    result_risk_major = extra_cloumn_info(indata,"risks")
    result_risk_minor = extra_cloumn_info(indata,"risks2")

    for _ in range(len(result_risk_major)) :
        tmp_result_double_dict = {}
        tmp_result_double_list = []
        tmp_result_max_double_ratio = 0
        tmp_result_min_double_sum = 1
        tmp_risk_major = result_risk_major[_]
        tmp_risk_minor = result_risk_minor[_]
        for _pos in range(0,len(tmp_risk_major),2) :
            tmp_pos = re.sub(u"\D","",tmp_risk_major[_pos])
            tmp_result_double_dict[tmp_pos] = float(tmp_risk_major[_pos+1])
        for _pos in range(0,len(tmp_risk_minor),2) :
            tmp_pos_info = tmp_risk_minor[_pos]
            tmp_minor_prob = tmp_risk_minor[_pos+1]
            tmp_pos = re.sub(u"\D","",tmp_risk_minor[_pos])
            if tmp_pos in tmp_result_double_dict.keys() :
                if tmp_result_double_dict[tmp_pos] != 0 :
                    tmp_ratio = float(tmp_minor_prob) / tmp_result_double_dict[tmp_pos]
                    tmp_ratio_format = '{:.2%}'.format(tmp_ratio)
                else :
                    tmp_ratio = float(0)
                tmp_sum = float(tmp_minor_prob) + tmp_result_double_dict[tmp_pos]
                if tmp_ratio > tmp_result_max_double_ratio :
                    tmp_result_max_double_ratio = tmp_ratio
                if tmp_sum < tmp_result_min_double_sum :
                    tmp_result_min_double_sum = tmp_sum
                tmp_output_list = [tmp_pos_info,tmp_result_double_dict[tmp_pos],tmp_minor_prob,tmp_ratio_format]      
                tmp_result_double_list.append(tmp_output_list)
        result_double_list.append(tmp_result_double_list)
        result_max_double_ratio.append(tmp_result_max_double_ratio)
        result_min_double_sum.append(tmp_result_min_double_sum)

    return result_double_list,result_max_double_ratio,result_min_double_sum


def get_wellinfo(infile) :
    """
    infile format :
    Total_wells,Well=wellType
    DWP-01-2022-10-30_1_A,DWP-01-2022-10-30_10_C=XF1
    ......
    """
    well_total  = []
    well_draw_dict = {}

    if isinstance(infile,str) :
        with open(infile) as f :
            for line in f.readlines():
                line = line.strip()
                _list = line.split(",")
                well_total.append(_list[0])
                if len(_list) == 2 :
                    t_well,t_type = _list[1].split("=")[0],_list[1].split("=")[1]
                    if t_type in well_draw_dict.keys() :
                        well_draw_dict[t_type].append(t_well)
                    else :
                        well_draw_dict[t_type] = [t_well]

    elif isinstance(infile,pd.DataFrame):
        for index, row in infile.iterrows():
            well_total.append(row['well'])
            if row['Nanopore_Result'] :
                t_well,t_type = row['Nanopore_Result'].split("=")[0], row['Nanopore_Result'].split("=")[1]
                if t_type in well_draw_dict.keys() :
                    well_draw_dict[t_type].append(t_well)
                else :
                    well_draw_dict[t_type] = [t_well]
    else:
        well_total, well_draw_dict = infile

    return well_total,well_draw_dict


def get_well_draw_dict(indata: pd.DataFrame) :
    well_draw_dict = defaultdict(list)
    for _, row in indata.iterrows():
        if row['Nanopore_Result'] and row['Nanopore_Result']!="Wrong":
            well_draw_dict[row['Nanopore_Result']].append(row['well'])
    return well_draw_dict


def get_sheetname_formwell(wellinfo_list) :
    tmp_sheet_name = []
    for i in sorted(wellinfo_list) :
        tmpi = i.split("-")
        tmpend = tmpi[-1].split("_")
        tmpwell = "-".join(tmpi[:-1]) + "-" + tmpend[0]
        tmp_sheet_name.append(tmpwell)
    sheet_name = [i for i in sorted(set(tmp_sheet_name))]
    return sheet_name

def draw_96_wellplate_empty(workbook,sheet_name) :
    """
    Draw in (0,0)
    """
    custom_font = Font(name='Light', size=12, bold='u')
    border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )

    custom_row = string.ascii_uppercase
    well_row = ["A","B","C","D","E","F","G","H"]
    well_col = [i for i in range(1,13)]

    workbook[sheet_name].cell(row=1,column=1,value="编号:"+sheet_name).font = custom_font
    workbook[sheet_name].cell(row=1,column=1,value="编号:"+sheet_name).border= border 
    workbook[sheet_name].cell(row=1,column=1,value="编号:"+sheet_name).alignment = Alignment(horizontal='center', vertical='center')
    workbook[sheet_name].merge_cells('A1:M1')
    for tmpi in range(1,14):
        workbook[sheet_name].row_dimensions[tmpi].height = 30
        for tmpj in range(1,11) :
            workbook[sheet_name].cell(row=tmpj,column=tmpi).border=border
    for tmpj in custom_row[:13]:
        workbook[sheet_name].column_dimensions[tmpj].width = 6
    for t_r in range(len(well_row)) :
        tmprow = well_row[t_r]
        workbook[sheet_name].cell(row=t_r+3,column=1,value=tmprow).font = custom_font
        workbook[sheet_name].cell(row=t_r+3,column=1,value=tmprow).alignment = Alignment(horizontal='center', vertical='center')
    for t_c in range(len(well_col)) :
        tmpcol = well_col[t_c]
        workbook[sheet_name].cell(row=2,column=t_c+2,value=tmpcol).font = custom_font
        workbook[sheet_name].cell(row=2,column=t_c+2,value=tmpcol).alignment = Alignment(horizontal='center', vertical='center')

def draw_96_wellplate_well(workbook,welltype : str,well : list,welltype_fill : dict, ref:pd.DataFrame) :
    """
    well_right = ["DWP-01-","DWP-01-",...]

    Right : T => green_fill 
            Double_Peak => dark_green_fill
    XF    : XF[num] => yellow_fill  :: e.g. XF1 XF2
    Wrong : F => ref_fill
    Unsure: U => grey_fill
    """
    
    custom_well_info = {}
    custom_font2 = Font(name='Calibri', size=12, bold='u')
    
    for i in well :
        twell = i.split("_")
        twell_name = twell[0]
        twell_col = int(twell[1]) + 1
        twell_row = ord(twell[2]) - 65 + 3 
        row = twell[1] if len(twell[1]) == 2 else '0' + twell[1]
        depth = ref[(ref['Plate']==twell_name) & (ref['Well_Row']==twell[2]) & (ref['Well_column']==row)]['mean_depth'].values[0]
        depth = depth if depth > 0 else 0
        if twell_name not in custom_well_info.keys() :
            custom_well_info[twell_name] = [twell[1:3]]
        else :
            custom_well_info[twell_name].append(twell[1:3])
        workbook[twell_name].cell(row=twell_row ,column=twell_col,value= f'{welltype_fill[welltype][0]}{chr(10)}{int(depth+0.999)}').font = custom_font2
        workbook[twell_name].cell(row=twell_row ,column=twell_col).fill = welltype_fill[welltype][1]
        workbook[twell_name].cell(row=twell_row ,column=twell_col).alignment = Alignment(horizontal='center', wrap_text=True)
    
    return custom_well_info

def fill_wellinfo(workbook,well_type,flag_offset,well_info,welltype_fill : dict) :
    for t_key in well_info.keys():
        for t_all in well_info[t_key]:
            tmp_row = flag_offset[t_key]["_".join(t_all)]
            workbook[t_key].cell(row=tmp_row, column=17).fill = welltype_fill[well_type][1]

def fill_whole_wrong_gene(workbook,sheet_name) :  
    for i in sheet_name :
        tmp_col_index = 1
        for tmp_col in range(2,14) :
            tmp_list = []
            for tmp_row in range(3,11) :
                r = workbook[i].cell(row=tmp_row,column=tmp_col).value
                tmp_list.append(r.split()[0] if r else r)
            if None not in tmp_list :
                tmp_col_index = tmp_col_index + 1
            if "T" not in tmp_list and "DP" not in tmp_list :
                if None not in tmp_list:
                    workbook[i].cell(row=tmp_col_index + 10,column=5).fill = PatternFill("solid",fgColor='F54937')
                    workbook[i].cell(row=tmp_col_index + 10,column=6).fill = PatternFill("solid",fgColor='F54937')
            right_num = tmp_list.count("T") + tmp_list.count("DP")
            if right_num == 1 :
                workbook[i].cell(row=tmp_col_index + 10,column=5).fill = PatternFill("solid",fgColor='F58585')
                workbook[i].cell(row=tmp_col_index + 10,column=6).fill = PatternFill("solid",fgColor='FF8585')


def draw_wellplate_sample_info_with_df(workbook,sheet_name,well_info_DF) :
    ref = well_info_DF[well_info_DF['Plate']==sheet_name].sort_values(by=['Well_column', 'Well_Row'])
    ref['Well-ID'] = ref['Well_Row'] + ref['Well_column'].astype(str)
    ref = ref.reset_index(drop=True)

    result_filled_clone_pos = {}
    custom_font3 = Font(name='Light', size=12)
    start_column = 16
    columns = ['Well-ID','clone_id', 'mean_depth', 'variants', 'Nanopore_Result', 'Choosed']
    for i, column_name in enumerate(columns):
        workbook[sheet_name].cell(row=1,column=start_column+i,value=column_name).font = custom_font3
        workbook[sheet_name].cell(row=1,column=start_column+i).alignment = Alignment(horizontal='center', vertical='center')

    workbook[sheet_name].cell(row=1,column=16,value='Well-ID').font = custom_font3
    workbook[sheet_name].cell(row=1,column=16).alignment = Alignment(horizontal='center', vertical='center')
    workbook[sheet_name].cell(row=1,column=17,value='clone_id').font = custom_font3
    workbook[sheet_name].cell(row=1,column=17).alignment = Alignment(horizontal='center', vertical='center')

    # fill all_info to write csv
    a_index = 0
    for index, row in ref.iterrows():
        result_filled_clone_pos['_'.join([str(int(row['Well_column'])), row['Well_Row'],])] = index+2
        for i, column_name in enumerate(columns):
            workbook[sheet_name].cell(row=index+2,column=start_column+i,value=row[column_name]).font = custom_font3
            workbook[sheet_name].cell(row=index+2,column=start_column+i).alignment = Alignment(horizontal='center', vertical='center')

        if row['Well_Row'] == 'A':
            workbook[sheet_name].cell(row=a_index+12, column=5, value=row['Well-ID']).font = custom_font3
            workbook[sheet_name].cell(row=a_index+12, column=5).alignment = Alignment(horizontal='center', vertical='center')
            workbook[sheet_name].cell(row=a_index+12, column=6, value=row['gene_id']).font = custom_font3
            workbook[sheet_name].cell(row=a_index+12, column=6).alignment = Alignment(horizontal='left', vertical='center')
            a_index += 1
        
    
    return result_filled_clone_pos

def agg_mutations(row):
    result = []
    for col in 'mutations	nucleotides	indels	deletions'.split():
        seg = str(row[col]).split("'")
        if len(seg) > 1:
            for i in range(1, len(seg), 2):
                result.append(seg[i])
    return str(list(set(result)))
    

def extract_well_info(well_info):
    well_info = well_info.split('_')
    alpha = well_info[-1]
    num = well_info[-2] if len(well_info[-2]) == 2 else '0' + well_info[-2]
    plate = '_'.join(well_info[:-2])
    return plate, alpha, num


def dictMerge(dict1, dict2): 
    res = {**dict1, **dict2} 
    return res 

def get_plasmid_primer_from_xlsx(infile : str) -> dict:
    wb = load_workbook(infile)
    ws = wb[wb.sheetnames[0]]

    plasmid_primer = {}

    plasmid_name_colindex = 0
    primer_info_colindex = 0
    tmp_index = 0
    for col in ws[1] :
        if col.value == "质粒名称" :
            plasmid_name_colindex = chr(tmp_index+97)
        if col.value == "测序引物" :
            primer_info_colindex = chr(tmp_index+97)
        tmp_index = tmp_index + 1

    tmp_index = 1
    for row in ws[plasmid_name_colindex] :
        if tmp_index > 1 :
            if row.value :
                plasmid_primer[row.value] = ws[primer_info_colindex+str(tmp_index)].value
        tmp_index = tmp_index + 1

    return plasmid_primer

def get_plasmid_primer(indir : str) -> dict :
    plasmid_primer = {}
    for tmpdir in os.listdir(indir) :
        tmpfile = indir + "/" + tmpdir
        plasmid_primer = dictMerge(plasmid_primer,get_plasmid_primer_from_xlsx(tmpfile))
    return plasmid_primer

def get_plasmid_primer_from_filled_xlsx(infile : str) -> dict :
    wb = load_workbook(infile)
    sheetsname = wb.sheetnames

    plasmid_primer = {}
    for shtname in sheetsname :
        ws = wb[shtname]
        tmp_plasmid_primer = {}
        for row in range(1,16) :
            #if ws.cell(row,3).value :
            if ws.cell(row,5).value :
                tmp_geneid = ws.cell(row,5).value.split(" ")[0]
                tmp_primer = ws.cell(row,8).value
                tmp_plasmid_primer[tmp_geneid] = tmp_primer
        plasmid_primer = dictMerge(plasmid_primer,tmp_plasmid_primer)

    return plasmid_primer

def comparsion_id(target,query) -> bool:
    '''
        Record special case
    '''
    flag = False

    target_info = re.split("[-_ ]",target.upper())
    query_info = re.split("[-_ ]",query.upper())

    tmp_index = 0
    for i in query_info :
        if i in target_info :
            tmp_index = tmp_index + 1
    if tmp_index >= min(len(target_info),len(query_info)) :
        flag = True
    
    return flag

def copy_template_xltx(template : str, result_wb : Workbook, result_sheetname : list) -> Workbook:
    template_wb = load_workbook(template)
    template_sheetname = template_wb.sheetnames
    template_wb_ws = template_wb[template_sheetname[0]]

    for i in result_sheetname :
        if i not in result_wb.sheetnames :
            result_wb.create_sheet(i)
        for row in template_wb_ws:
            for cell in row:
                result_wb[i][cell.coordinate].value = cell.value
                if cell.has_style:
                    result_wb[i][cell.coordinate].font = copy(cell.font)
                    result_wb[i][cell.coordinate].border = copy(cell.border)
                    result_wb[i][cell.coordinate].fill = copy(cell.fill)
                    result_wb[i][cell.coordinate].number_format = copy(cell.number_format)
                    result_wb[i][cell.coordinate].protection = copy(cell.protection)
                    result_wb[i][cell.coordinate].alignment = copy(cell.alignment)
        for tmp_row in range(1, template_wb_ws.max_row + 1):
            result_wb[i].row_dimensions[tmp_row].height = template_wb_ws.row_dimensions[tmp_row].height
        for tmp_col in range(1, template_wb_ws.max_column + 1):
            result_wb[i].column_dimensions[get_column_letter(tmp_col)].width = template_wb_ws.column_dimensions[get_column_letter(tmp_col)].width

    return result_wb

def writexlsx(template : str, refprimer : dict, nanopore_result : str) :
    nanopore_wb = load_workbook(nanopore_result)
    sheetnames = nanopore_wb.sheetnames
    result_wb = Workbook()

    result_wb = copy_template_xltx(template,result_wb,sheetnames)
    result_wb.remove(result_wb["Sheet"])

    for i in sheetnames :
        ws = nanopore_wb[i]
        result_ws = result_wb[i]
        tmp_geneinfo = {}
        for row in range(1,97) :
            if ws.cell(row,16).value :
                tmp_plasmid_wellid = ws.cell(row,16).value
                tmp_plasmid_name = ws.cell(row,17).value
                tmp_plasmid_name_gene = "-".join(ws.cell(row,17).value.split("-")[:-1])
                tmp_cell_fillstyle = copy(ws.cell(row,17).fill)
                tmp_geneinfo[tmp_plasmid_wellid] = [tmp_plasmid_name,tmp_cell_fillstyle,tmp_plasmid_name_gene]
        for row in range(3,100) :
            tmp_well_index = result_ws.cell(row,1).value
            if tmp_well_index in tmp_geneinfo.keys() :
                result_ws.cell(row,2).value = tmp_geneinfo[tmp_well_index][0]
                result_ws.cell(row,2).fill = tmp_geneinfo[tmp_well_index][1]
                for key in refprimer.keys() :
                    #if comparsion_id(tmp_geneinfo[tmp_well_index][2],key) :
                    if comparsion_id(key,tmp_geneinfo[tmp_well_index][2]) :
                        result_ws.cell(row,7).value = refprimer[key]
        # Merge_Primer_cells
        for row in range(3,99,8) :
            result_ws.merge_cells(start_row=row,end_row=row+7,start_column=7,end_column=7) 

    outfile = ".".join(nanopore_result.split(".")[:-1]) + "_add_primer.xlsx"
    result_wb.save(outfile)


def formate_excel() :
    font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False)
    alignment=Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
    fill = PatternFill("solid", fgColor="FFFF00")

    return font,alignment,fill

def real_plasmid_name(name : str) -> list :
    prefix = re.split(r'[-_]',name)[:2]
    if len(prefix) < 2 :
        return ""
    len_prefix = len("-".join(prefix)) + 1
    plasmid_name = name[len_prefix:]
    return plasmid_name

def filter_result(indata : pd.DataFrame) -> pd.DataFrame :
    keep_type = ["N","CL","JC"]
    result = indata[indata["CL/XF/N/Y"].str.contains("XF",case=False) == False]
    result = result[result["CL/XF/N/Y"].str.contains("Y",case=False) == False]
    return result


def Result_Filter_add_XF_add_risk_unsure(infile_bwa:pd.DataFrame, infile_medaka:pd.DataFrame, infile_barcode:pd.DataFrame,):
    """
    07.Result_Filter_add_XF_add_risk_unsure.py
    """
    infile_bwa["filter_type"] = "Wrong"
    infile_bwa.loc[(infile_bwa["nucleotides"] == "[]") & (infile_bwa["barcode_type"]=="sum") & (infile_bwa["identity"] > 0.95) & (infile_bwa["mutations"] == "[]") & (infile_bwa["indels"] == "[]") & (infile_bwa["deletions"] == "[]") & (infile_bwa["mapped_reads"] >= 100),"filter_type"]="Check"
    infile_bwa.loc[(infile_bwa["mapped_reads"] < 100),"filter_type"]="Unsure"

    infile_medaka["filter_type"] = "Wrong"
    infile_medaka.loc[(infile_medaka["identity"] > 0.95) & (infile_medaka["variants"] == "[]"),"filter_type"] = "Check"

    indata = pd.merge(infile_bwa,infile_medaka,on=["well"],how="outer",suffixes=["_bwa","_medaka"])
    indata = indata.loc[(indata["barcode_type_bwa"]=="sum") & (indata["barcode_type_medaka"] == "sum")]
    indata = indata.drop_duplicates()

    # XF info
    tmp_xf, choose_xf = add_column_mutation_list(indata)
    # Double peaks
    output_double, output_double_ratio, output_double_sum_prob = add_column_Double_Peak(indata)

    indata["mutation_list"] = tmp_xf
    indata["mutation_set_num"] = np.array(choose_xf)
    indata["double_ratio_list"] = output_double
    indata["double_ratio_max"] = output_double_ratio
    indata["double_sum_prob_min"] = output_double_sum_prob

    indata["XF"] = "Wrong"
    indata.loc[(indata["mutation_set_num"] > 0) & (indata["mutation_set_num"] < 4),"XF"] = "Check"

    #indata["Nanopore_Result"] = "XF" + indata.loc[(indata["XF"] == "Check") & (indata["mapped_reads"] >= 50),"mutation_set_num"].astype(str)
    indata["Nanopore_Result"] = "XF" + indata.loc[(indata["XF"] == "Check") & (indata["mapped_reads"] >= 100),"mutation_set_num"].astype(str)

    #indata.loc[((indata["filter_type_bwa"] == "Unsure") & (indata["XF"] == "Wrong")),"Nanopore_Result"] = "Unsure"
    indata.loc[(indata["filter_type_bwa"] == "Unsure"),"Nanopore_Result"] = "Unsure"

    indata.loc[(indata["filter_type_bwa"] == "Check") & (indata["filter_type_medaka"] == "Check"),"Nanopore_Result"] = "Right"
    # indata.loc[((indata["double_ratio_max"] >= 0.40) | (indata["double_sum_prob_min"] <= 0.8)) & (indata["Nanopore_Result"] == "Right"),"Nanopore_Result"] = "Double_Peak"
    indata.loc[(indata["double_ratio_max"] >= 0.45) & (indata["Nanopore_Result"] == "Right"),"Nanopore_Result"] = "Double_Peak"

    indata.loc[indata["Nanopore_Result"].isnull(),"Nanopore_Result"] = "Wrong"
    

    well_total = infile_barcode["well"].to_list()
    well_draw_raw = np.array(indata.loc[indata["Nanopore_Result"] != "Wrong",["well","Nanopore_Result"]].sort_values("well").drop_duplicates())

    #with open("./Dev_NanoporeResultFilter/Result.csv","w") as r:
    Result = []

    for i in range(len(well_total)) :
        tmp_result = []
        tmp_result.append(well_total[i])
        if i < len(well_draw_raw) :
            tmp_result.append("=".join(well_draw_raw[i]))
        Result.append(tmp_result)
    
    Result = pd.DataFrame(Result,columns=["well","Nanopore_Result"])

    return indata, Result


def Result_Merge_for_Sanger(indata, well_reference_sample=None):
    if well_reference_sample is None:
        result = indata
    else:
        ref_pd = exact_clone_well_info(well_reference_sample) if isinstance(well_reference_sample, str) else well_reference_sample
        result = pd.merge(ref_pd,indata, how='left')

    result = result.sort_values("clone_id")
    result = result.drop_duplicates()
    result["Date"] = str(datetime.date.today())
    

    new_header = ["Date", "Nanopore_Result", "Choosed"]
    for col in result.columns:
        if col not in new_header:
            new_header.append(col)

    nanopore_info = result[new_header]
    return nanopore_info


def Result_Vis_add_risk_unsure_v2(Result, nanopore_info, out_dir='./'):
    welltype_fill = {
        "Right" : ["T",PatternFill("solid",fgColor='92D050')] ,
        "Double_Peak": ["DP",PatternFill("solid",fgColor='769345')] ,
        "Wrong" : ["F",PatternFill("solid",fgColor='F54937')] ,
        "XF1": ["XF1",PatternFill("solid",fgColor='FFF2CC')] ,
        "XF2": ["XF2",PatternFill("solid",fgColor='FFF2CC')] ,
        "XF3": ["XF3",PatternFill("solid",fgColor='FFF2CC')] ,
        "Unsure" : ["U",PatternFill("solid",fgColor="AEAAAA")]
    }
    ref = nanopore_info
    
    ref[['Plate', 'Well_Row', "Well_column"]] = ref['well'].apply(lambda x: pd.Series(extract_well_info(x)))

    well_total,well_draw_dict = get_wellinfo(Result)
    sheet_name = get_sheetname_formwell(well_total)
    dest_filename = "Nanopore_"+"-".join(sheet_name[0].split("-")[2:]) + ".xlsx"

    # ref_excel = openpyxl.load_workbook(reffile)
    wb = Workbook()
    flag_offset = {}
    for i in sheet_name :
        wb.create_sheet(title=i)
        wb[i].column_dimensions['Q'].width = 30
        wb[i].column_dimensions['R'].width = 30
        wb[i].column_dimensions['S'].width = 30
        wb[i].column_dimensions['T'].width = 30
        wb[i].column_dimensions['U'].width = 30
        wb[i].column_dimensions['V'].width = 60
        draw_96_wellplate_empty(wb,i)
        flag_offset[i] = draw_wellplate_sample_info_with_df(wb,i,ref)
        # flag_offset[i] = draw_wellplate_sample_info(wb,i,ref_well)

    wb.remove(wb["Sheet"])
    draw_96_wellplate_well(wb,"Wrong",well_total,welltype_fill, ref)
    for _type in well_draw_dict.keys() :
        if _type in ["Right","Double_Peak","XF1"] :
            tmp_plate_well = draw_96_wellplate_well(wb,_type,well_draw_dict[_type],welltype_fill, ref)
            fill_wellinfo(wb,_type,flag_offset,tmp_plate_well,welltype_fill)
        else :
            draw_96_wellplate_well(wb,_type,well_draw_dict[_type],welltype_fill, ref)

    fill_whole_wrong_gene(wb,sheet_name)
    filename = os.path.join(out_dir, dest_filename)
    wb.save(filename=filename)
    return filename


def Auto_write_from(nanopore_result_file, plasmid_ref='ref_primer.xlsx', template_file='Nanopore_Template.xltx'):
    plasmid_primer = get_plasmid_primer_from_filled_xlsx(plasmid_ref)
    writexlsx(template_file,plasmid_primer,nanopore_result_file)


def Stat_NanoporeResult(nanopore_info, outfile=None, suffix='') :
    data = nanopore_info 
    _deal_data = data[["clone_id","Nanopore_Result"]]
    
    _deal_data["质粒名_1"] = _deal_data["clone_id"].apply(lambda x: "-".join(x.split("-")[:-1]))
    _deal_data["感受态细胞"] = _deal_data["质粒名_1"].apply(lambda x: re.split(r'[-_]',x)[0])
    _deal_data["CL/XF/N/Y"] = _deal_data["质粒名_1"].apply(lambda x: re.split(r'[-_]',x)[1])
    _deal_data["Nanopore正确率_1"] = _deal_data.groupby(["质粒名_1"])["Nanopore_Result"].transform(lambda x: (x=='Right').sum())
    _deal_data["Nanopore正确率_2"] = _deal_data.groupby(["质粒名_1"])["Nanopore_Result"].transform(lambda x: (x=='Double_Peak').sum())
    _deal_data["Nanopore正确率_3"] = _deal_data["Nanopore正确率_1"] + _deal_data["Nanopore正确率_2"]
    
    result = _deal_data.groupby(["质粒名_1","Nanopore正确率_1","Nanopore正确率_3","感受态细胞","CL/XF/N/Y"]).count().reset_index()
    result["Nanopore正确率(不包含双峰)"] = result["Nanopore正确率_1"].astype(str) + "/" + result["clone_id"].astype(str)
    result["Nanopore正确率(包含双峰)"] = result["Nanopore正确率_3"].astype(str) + "/" + result["clone_id"].astype(str)
    result[["长度","长度区间"]] = ""
    _tmp = result["质粒名_1"].values.tolist()
    _name = []
    for _ in _tmp :
        _t = real_plasmid_name(_)
        if _t[-1] in ['-','_'] :
            _name.append(_t[:-1])
        else :
            _name.append(_t)
    result["质粒名"] = _name

    final = result[["质粒名","Nanopore正确率(不包含双峰)","长度","长度区间","感受态细胞","CL/XF/N/Y"]]

    final = filter_result(final)
    
    final["引物合成时间"] = ""
    final["纠错酶类型"] = ""
    final["Source"] = "Nanopore_" + suffix
    # final.to_excel(outfile,index=None)

    if outfile is None:
        return final

    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    font,alignment,fill = formate_excel()
    for _row in range(final.shape[0]) :
        for _col in range(final.shape[1]) :
            ws.cell(_row + 2,_col + 1).value = final.iloc[_row,_col]
            ws.cell(_row + 2,_col + 1).font = font
            ws.cell(_row + 2,_col + 1).alignment = alignment

    header = ["质粒名","正确率","长度","长度区间","感受态细胞","CL/XF/N/Y","引物合成时间","纠错酶类型","Source"]
    for _ in range(len(header)) :
        ws.cell(1,_+1).value = header[_]
        ws.cell(1,_ + 1).font = font
        ws.cell(1,_ + 1).alignment = alignment

    ws.column_dimensions["A"].width= 45
    ws.column_dimensions["D"].width= 12
    ws.column_dimensions["E"].width= 12
    ws.column_dimensions["G"].width= 12
    ws.column_dimensions["H"].width= 12
    ws.column_dimensions["I"].width= 30
    wb.save(outfile)


def trans_list_column_to_str(df):
    for col in df.columns:
        if isinstance(df[col].iloc[0], list):
            df[col] = df[col].apply(str)
    return df


def mutation_set_count(xf):
    result = 0
    xf_length = 30
    start = - xf_length
    for m in xf:
        i = 0
        r = m['Reference']
        while i < len(r):
            pos = m['Position'] + i
            if pos - start >= xf_length:
                result += 1
                start = pos
                i += 1
            else:
                i += start - pos + xf_length 
    return result


def nanopore_result_type(row):
    xf = row['variants']
    xf = sorted(xf, key=lambda x: x['Position'])
    xf_count = mutation_set_count(xf)
    if xf_count:
        xf_count = mutation_set_count(xf+[m for m in row['DP'] if m['Ratio'] > 0.75])
    if xf_count > 3 or row['identity'] < (0.999):
        return 'Wrong'
    
    if pd.isna(row['mean_depth']) or row['mean_depth'] < 30 or row['match_rate'] < 0.1 or row['match_rate'] < 0.5 and xf_count:
        return 'Unsure'
    
    if xf_count:
        return f"XF{xf_count}"
    
    if len(row['DP']) > 0:
        return 'Double_Peak'
    
    return 'Right'


def risk_score(mutations):
    scores = [m['Ratio'] * min(m['Plus_Ratio'], 1-m['Plus_Ratio']) / 0.5 * (1+0.3*(len(m['Reference'])>2))
              for m in mutations] + [0]
    return max(scores)


def format_mutation(mutations, sort_by_ratio=False):
    res = []
    if sort_by_ratio:
        mutations = sorted(mutations, key=lambda x: x['Ratio'], reverse=True)
    for m in mutations:
        res.append(f"{m['Reference']}{m['Position']}{m['Mutation']}:{m['Ratio']:.3f}:{m['Plus_Ratio']:.3f}")

    return '; '.join(res)


def add_choose_result(df:pd.DataFrame)->pd.DataFrame:
    result = []
    choose_count = 2

    df['Nanopore_Result'] = df.apply(nanopore_result_type, axis=1)
    df['DP_score'] = df['DP'].apply(risk_score)
    df['Risk_score'] = df['Risk'].apply(risk_score)

    for i, group in df.groupby('known_seq_id'):
        right_filter = group['Nanopore_Result'] == 'Right'
        right_count = sum(right_filter)
        double_peak_filter = group['Nanopore_Result'] == 'Double_Peak'
        double_peak_count = sum(double_peak_filter)

        group['Right'] = f"{right_count}/{len(group)}"
        group['Right & DP'] = f"{right_count + double_peak_count}/{len(group)}" 
        group['Choosed'] = ''
        
        if right_count+double_peak_count:
            right = group[right_filter | double_peak_filter] if right_count < choose_count else group[right_filter]
            right = right[['DP_score', 'Risk_score', 'match_rate', 'mean_depth']]
            right['DP_score'] = right['DP_score'] // 0.05
            right['Risk_score'] = right['Risk_score'] // 0.05
            right['match_rate_'] = right['match_rate'] // 0.05
            right = right.sort_values(['DP_score', 'Risk_score', 'match_rate_', 'mean_depth'], ascending=[True, True, False, False])
            group.loc[right.head(choose_count).index, 'Choosed'] = 'Choose'

        result.append(group)

    result = pd.concat(result).drop(['DP_score', 'Risk_score'], axis=1)

    result['variants'] = result['variants'].apply(format_mutation)
    result['DP'] = result['DP'].apply(format_mutation)
    result['Risk'] = result['Risk'].apply(lambda x: format_mutation(x, sort_by_ratio=True))

    return result


if __name__=='__main__':
    input_file = '' #'03.Analysis/N2025-02-26-520-4h/sva_result/summary.csv'
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]

    nanopore_info = pd.read_csv(input_file)
    nanopore_info['variants'] = nanopore_info['variants'].apply(lambda x: json.loads(x.replace("'", '"').replace('inf', '1')))
    nanopore_info['DP'] = nanopore_info['DP'].apply(lambda x: json.loads(x.replace("'", '"').replace('inf', '1')))
    nanopore_info['Risk'] = nanopore_info['Risk'].apply(lambda x: json.loads(x.replace("'", '"').replace('inf', '1')))
    nanopore_info = add_choose_result(nanopore_info)
    Result = (nanopore_info.well.tolist(), get_well_draw_dict(nanopore_info))

    nanopore_info = trans_list_column_to_str(nanopore_info)
    nanopore_info = Result_Merge_for_Sanger(nanopore_info)#, 'well_reference_sample_2024-10-21.xlsx')
    file_name = Result_Vis_add_risk_unsure_v2(Result, nanopore_info)
    nanopore_info.to_csv(file_name.replace('.xlsx', '_summary.csv'), index=False)
    Stat_NanoporeResult(nanopore_info, file_name.replace('Nanopore_', 'Nanopore_Stat_'), file_name.split('_')[-1].split('.')[0])



    pass
